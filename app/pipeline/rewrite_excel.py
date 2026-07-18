"""Rewrite an Excel workbook's raw bytes, passing every leaf string cell
value through a caller-supplied callback while preserving the workbook's
structure (sheets, rows, columns) and, for modern formats, existing
formatting and formulas on cells the callback never sees.

This module makes no decisions about what to redact: the ``transform``
callback IS the redaction decision, already made by the caller. This module
only shuttles cell text through it and writes the result back out.

Legacy binary ``.xls`` is the one format this module cannot re-emit as-is:
modern Python has no maintained writer for it (xlwt is unmaintained and
does not round-trip modern files reliably), so ``.xls`` input is read with
xlrd and re-emitted as a brand-new ``.xlsx`` workbook instead — i.e. this
module silently upgrades legacy .xls to .xlsx on output. Call
``output_suffix_for(suffix)`` to learn the suffix the output bytes actually
have before naming the output file; don't assume it always matches the
input suffix.
"""

import io
from typing import Callable

import openpyxl
import xlrd

_XLSX_SUFFIXES = (".xlsx", ".xlsm", ".xltx", ".xltm")


def output_suffix_for(suffix: str) -> str:
    """Return the file suffix the rewritten bytes should be saved with.

    Every modern Excel suffix round-trips as itself. Legacy ".xls" is the one
    exception: rewrite_excel() has no way to write the legacy binary format
    back out (see its docstring), so it upgrades the output to ".xlsx". This
    helper lets a caller pick the right output filename extension without
    having to know that upgrade rule itself.
    """
    if suffix.lower() == ".xls":
        return ".xlsx"
    return suffix.lower()


def _is_formula(value: str) -> bool:
    # Without data_only=True, openpyxl surfaces formula cells as their
    # formula text (e.g. "=SUM(A1:A2)") — a str, but not leaf content we
    # should hand to the callback: rewriting it would corrupt the formula.
    return value.startswith("=")


def _rewrite_xlsx(original_bytes: bytes, transform: Callable[[str], str]) -> bytes:
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(original_bytes))
    except Exception as exc:
        raise ValueError(f"could not be opened as an Excel file: {exc}") from exc

    for sheet in workbook.worksheets:
        # Row 1 is left untouched, on the assumption (true for virtually
        # every real spreadsheet this app sees) that it's a header row, not
        # data — an isolated header word like "Vorname" or "E-Mail" has no
        # sentence context, and spaCy's NER can misclassify it as PII in
        # exactly the way it does real short-and-context-free data cells (a
        # header actually getting relabeled "[PERSON47]" was observed),
        # scrambling column labels in the reusable copy for no privacy
        # benefit — a header is metadata describing the column, not user
        # data. See app/pipeline/column_classifier.py, which relies on
        # headers staying legible to classify columns in the first place.
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                value = cell.value
                if isinstance(value, str) and not _is_formula(value):
                    cell.value = transform(value)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


_INVALID_SHEET_CHARS = set('[]:*?/\\')


def _safe_sheet_title(name: str, used: set[str]) -> str:
    """Sanitize an xlrd sheet name so openpyxl accepts it as a worksheet
    title: strip characters Excel forbids, cap the 31-char length limit, and
    disambiguate collisions that truncation or sanitizing might create.
    """
    cleaned = "".join(ch for ch in (name or "Sheet") if ch not in _INVALID_SHEET_CHARS)
    cleaned = cleaned.strip() or "Sheet"
    title = cleaned[:31]

    suffix_num = 1
    while title in used:
        suffix_num += 1
        tag = f"_{suffix_num}"
        title = cleaned[: 31 - len(tag)] + tag

    return title


def _rewrite_xls(original_bytes: bytes, transform: Callable[[str], str]) -> bytes:
    """Read a legacy .xls workbook with xlrd and re-emit it as a NEW .xlsx
    workbook via openpyxl.

    Modern Python has no maintained writer for the legacy binary .xls format
    (xlwt is unmaintained and does not round-trip modern files reliably), so
    this function upgrades the output format to .xlsx. Callers should use
    output_suffix_for(".xls") (-> ".xlsx") to name the resulting file rather
    than assuming the output suffix always matches the input suffix.
    """
    try:
        source = xlrd.open_workbook(file_contents=original_bytes)
    except Exception as exc:
        raise ValueError(f"could not be opened as an Excel file: {exc}") from exc

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    used_titles: set[str] = set()
    for sheet in source.sheets():
        title = _safe_sheet_title(sheet.name, used_titles)
        used_titles.add(title)
        target = workbook.create_sheet(title=title)
        for row_idx in range(sheet.nrows):
            for col_idx, value in enumerate(sheet.row_values(row_idx)):
                if value == "":
                    # xlrd represents a genuinely blank cell as "" — leave it
                    # out rather than handing empty strings to the callback.
                    continue
                # Row 0 is left untouched — see _rewrite_xlsx()'s comment on
                # why the (assumed) header row is skipped entirely.
                if isinstance(value, str) and row_idx > 0:
                    value = transform(value)
                target.cell(row=row_idx + 1, column=col_idx + 1, value=value)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def rewrite_excel(
    original_bytes: bytes, suffix: str, transform: Callable[[str], str]
) -> bytes:
    """Return new Excel file bytes with every string cell value replaced by
    transform(value), preserving sheets, rows, columns, and (for the
    openpyxl path) formatting and untouched formulas.

    ".xls" input is a special case: it is upgraded to a plain .xlsx workbook
    on output, since there is no reliable modern writer for the legacy binary
    format. Use output_suffix_for(suffix) to get the correct output suffix.
    """
    suffix = suffix.lower()

    if suffix == ".xls":
        return _rewrite_xls(original_bytes, transform)

    if suffix in _XLSX_SUFFIXES:
        return _rewrite_xlsx(original_bytes, transform)

    raise ValueError(
        f"unsupported Excel suffix '{suffix}' for rewrite_excel()."
    )
